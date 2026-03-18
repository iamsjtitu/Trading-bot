"""
Capital Gains Tax Calculator - Indian Tax System
F&O Options Trading Tax Report Generator
"""
import io
import os
from datetime import datetime, timezone
from typing import Dict, List
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from fpdf import FPDF


# Indian Tax Constants
STCG_RATE = 0.15          # 15% Short Term Capital Gains on listed securities
CESS_RATE = 0.04          # 4% Health & Education Cess
SURCHARGE_10 = 0.10       # 10% surcharge (income 50L-1Cr)
SURCHARGE_15 = 0.15       # 15% surcharge (income > 1Cr)
STT_RATE_OPTIONS = 0.0625 # 0.0625% STT on options sell side
F_AND_O_AUDIT_LIMIT = 10_00_00_000  # 10 Crore turnover for tax audit (digital)


def get_fy_range(fy_year: str):
    """Get start and end dates for a Financial Year like '2025-26'"""
    parts = fy_year.split('-')
    start_year = int(parts[0])
    start = datetime(start_year, 4, 1, tzinfo=timezone.utc)
    end = datetime(start_year + 1, 3, 31, 23, 59, 59, tzinfo=timezone.utc)
    return start.isoformat(), end.isoformat()


def calculate_tax_report(trades: List[Dict], fy_year: str) -> Dict:
    """Calculate capital gains tax report for a financial year"""
    fy_start, fy_end = get_fy_range(fy_year)

    # Filter trades for the FY
    fy_trades = [t for t in trades if t.get('status') == 'CLOSED'
                 and (t.get('exit_time') or t.get('entry_time', '')) >= fy_start
                 and (t.get('exit_time') or t.get('entry_time', '')) <= fy_end]

    if not fy_trades:
        return _empty_report(fy_year)

    # Calculate totals
    total_buy_value = sum(t.get('investment', 0) for t in fy_trades)
    total_sell_value = sum((t.get('exit_price', 0) or 0) * (t.get('quantity', 0) or 0) for t in fy_trades)
    total_pnl = sum(t.get('pnl', 0) for t in fy_trades)
    profitable_trades = [t for t in fy_trades if (t.get('pnl') or 0) > 0]
    loss_trades = [t for t in fy_trades if (t.get('pnl') or 0) < 0]
    total_profit = sum(t.get('pnl', 0) for t in profitable_trades)
    total_loss = abs(sum(t.get('pnl', 0) for t in loss_trades))

    # F&O Turnover = sum of absolute P&L per trade (for audit purposes)
    turnover = sum(abs(t.get('pnl', 0)) for t in fy_trades)

    # STT paid (approximate - on sell side of options)
    stt_paid = total_sell_value * STT_RATE_OPTIONS / 100

    # Tax calculation
    net_profit = total_pnl
    stcg_tax = max(0, net_profit * STCG_RATE) if net_profit > 0 else 0
    cess = stcg_tax * CESS_RATE
    total_tax = stcg_tax + cess

    # Monthly breakdown
    monthly = {}
    for t in fy_trades:
        exit_dt = t.get('exit_time') or t.get('entry_time', '')
        if exit_dt:
            month_key = exit_dt[:7]  # YYYY-MM
            if month_key not in monthly:
                monthly[month_key] = {'trades': 0, 'profit': 0, 'loss': 0, 'net_pnl': 0, 'turnover': 0, 'buy_value': 0, 'sell_value': 0}
            m = monthly[month_key]
            m['trades'] += 1
            pnl = t.get('pnl', 0)
            m['net_pnl'] += pnl
            m['turnover'] += abs(pnl)
            m['buy_value'] += t.get('investment', 0)
            m['sell_value'] += (t.get('exit_price', 0) or 0) * (t.get('quantity', 0) or 0)
            if pnl > 0:
                m['profit'] += pnl
            else:
                m['loss'] += abs(pnl)

    # Add tax for each month
    for k, m in monthly.items():
        m['stcg_tax'] = round(max(0, m['net_pnl'] * STCG_RATE), 2) if m['net_pnl'] > 0 else 0
        m['cess'] = round(m['stcg_tax'] * CESS_RATE, 2)
        m['total_tax'] = round(m['stcg_tax'] + m['cess'], 2)

    # Sort monthly by key
    monthly_sorted = dict(sorted(monthly.items()))

    # Tax audit check
    audit_required = turnover > F_AND_O_AUDIT_LIMIT

    return {
        'fy_year': fy_year,
        'total_trades': len(fy_trades),
        'profitable_trades': len(profitable_trades),
        'loss_trades': len(loss_trades),
        'win_rate': round((len(profitable_trades) / max(len(fy_trades), 1)) * 100, 1),
        'total_buy_value': round(total_buy_value, 2),
        'total_sell_value': round(total_sell_value, 2),
        'total_profit': round(total_profit, 2),
        'total_loss': round(total_loss, 2),
        'net_pnl': round(net_profit, 2),
        'turnover': round(turnover, 2),
        'stt_paid': round(stt_paid, 2),
        'stcg_tax': round(stcg_tax, 2),
        'cess': round(cess, 2),
        'total_tax_liability': round(total_tax, 2),
        'effective_tax_rate': round((total_tax / max(net_profit, 1)) * 100, 1) if net_profit > 0 else 0,
        'audit_required': audit_required,
        'audit_limit': F_AND_O_AUDIT_LIMIT,
        'monthly_breakdown': monthly_sorted,
        'trades': fy_trades,
    }


def _empty_report(fy_year):
    return {
        'fy_year': fy_year, 'total_trades': 0, 'profitable_trades': 0, 'loss_trades': 0,
        'win_rate': 0, 'total_buy_value': 0, 'total_sell_value': 0, 'total_profit': 0,
        'total_loss': 0, 'net_pnl': 0, 'turnover': 0, 'stt_paid': 0, 'stcg_tax': 0,
        'cess': 0, 'total_tax_liability': 0, 'effective_tax_rate': 0, 'audit_required': False,
        'audit_limit': F_AND_O_AUDIT_LIMIT, 'monthly_breakdown': {}, 'trades': [],
    }


def generate_excel_report(report: Dict) -> bytes:
    """Generate Excel report with tax calculations"""
    wb = Workbook()

    # Styles
    header_font = Font(bold=True, size=14, color="FFFFFF")
    header_fill = PatternFill(start_color="1a365d", end_color="1a365d", fill_type="solid")
    sub_header = Font(bold=True, size=11)
    sub_fill = PatternFill(start_color="e2e8f0", end_color="e2e8f0", fill_type="solid")
    green_font = Font(color="16a34a", bold=True)
    red_font = Font(color="dc2626", bold=True)
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    currency_fmt = '#,##0.00'

    # ===== Sheet 1: Summary =====
    ws = wb.active
    ws.title = "Tax Summary"
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 25

    # Title
    ws.merge_cells('A1:B1')
    ws['A1'] = f"Capital Gains Tax Report - FY {report['fy_year']}"
    ws['A1'].font = header_font
    ws['A1'].fill = header_fill
    ws['A1'].alignment = Alignment(horizontal='center')
    ws['B1'].fill = header_fill

    row = 3
    summary_data = [
        ("TRADING SUMMARY", ""),
        ("Total Trades", report['total_trades']),
        ("Profitable Trades", report['profitable_trades']),
        ("Loss-Making Trades", report['loss_trades']),
        ("Win Rate", f"{report['win_rate']}%"),
        ("", ""),
        ("FINANCIALS", ""),
        ("Total Buy Value", report['total_buy_value']),
        ("Total Sell Value", report['total_sell_value']),
        ("Total Profit", report['total_profit']),
        ("Total Loss", report['total_loss']),
        ("Net P&L", report['net_pnl']),
        ("F&O Turnover", report['turnover']),
        ("STT Paid (Approx)", report['stt_paid']),
        ("", ""),
        ("TAX CALCULATION", ""),
        ("STCG Tax @15%", report['stcg_tax']),
        ("Health & Education Cess @4%", report['cess']),
        ("Total Tax Liability", report['total_tax_liability']),
        ("Effective Tax Rate", f"{report['effective_tax_rate']}%"),
        ("", ""),
        ("COMPLIANCE", ""),
        ("Tax Audit Required (44AB)", "Yes" if report['audit_required'] else "No"),
        ("Audit Threshold", report['audit_limit']),
    ]

    for label, value in summary_data:
        if label in ("TRADING SUMMARY", "FINANCIALS", "TAX CALCULATION", "COMPLIANCE"):
            ws.cell(row=row, column=1, value=label).font = sub_header
            ws.cell(row=row, column=1).fill = sub_fill
            ws.cell(row=row, column=2).fill = sub_fill
        else:
            ws.cell(row=row, column=1, value=label)
            cell = ws.cell(row=row, column=2, value=value)
            if isinstance(value, (int, float)) and label:
                cell.number_format = currency_fmt
                if 'Profit' in label or label == 'Net P&L':
                    cell.font = green_font if value >= 0 else red_font
                if label == 'Total Tax Liability':
                    cell.font = Font(bold=True, size=12, color="dc2626")
        ws.cell(row=row, column=1).border = border
        ws.cell(row=row, column=2).border = border
        row += 1

    # ===== Sheet 2: Monthly Breakdown =====
    ws2 = wb.create_sheet("Monthly Report")
    headers = ["Month", "Trades", "Profit", "Loss", "Net P&L", "Turnover", "STCG Tax", "Cess", "Total Tax"]
    for col, h in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.border = border
        ws2.column_dimensions[get_column_letter(col)].width = 18

    row = 2
    for month, data in report['monthly_breakdown'].items():
        vals = [month, data['trades'], data['profit'], data['loss'], data['net_pnl'], data['turnover'], data['stcg_tax'], data['cess'], data['total_tax']]
        for col, v in enumerate(vals, 1):
            cell = ws2.cell(row=row, column=col, value=v)
            cell.border = border
            if col >= 3:
                cell.number_format = currency_fmt
            if col == 5:
                cell.font = green_font if v >= 0 else red_font
        row += 1

    # Totals row
    total_vals = ["TOTAL", report['total_trades'], report['total_profit'], report['total_loss'], report['net_pnl'], report['turnover'], report['stcg_tax'], report['cess'], report['total_tax_liability']]
    for col, v in enumerate(total_vals, 1):
        cell = ws2.cell(row=row, column=col, value=v)
        cell.font = Font(bold=True)
        cell.fill = sub_fill
        cell.border = border
        if col >= 3:
            cell.number_format = currency_fmt

    # ===== Sheet 3: Trade Details =====
    ws3 = wb.create_sheet("Trade Details")
    trade_headers = ["#", "Type", "Symbol", "Entry Date", "Exit Date", "Entry Price", "Exit Price", "Qty", "Buy Value", "Sell Value", "P&L", "P&L %", "Exit Reason"]
    for col, h in enumerate(trade_headers, 1):
        cell = ws3.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.border = border
        ws3.column_dimensions[get_column_letter(col)].width = 15

    for i, t in enumerate(report['trades'], 1):
        sell_value = (t.get('exit_price', 0) or 0) * (t.get('quantity', 0) or 0)
        vals = [
            i, t.get('trade_type', ''), t.get('symbol', ''),
            (t.get('entry_time', '') or '')[:10], (t.get('exit_time', '') or '')[:10],
            t.get('entry_price', 0), t.get('exit_price', 0) or 0,
            t.get('quantity', 0), t.get('investment', 0), round(sell_value, 2),
            t.get('pnl', 0), t.get('pnl_percentage', 0), t.get('exit_reason', ''),
        ]
        for col, v in enumerate(vals, 1):
            cell = ws3.cell(row=i + 1, column=col, value=v)
            cell.border = border
            if col in (6, 7, 9, 10, 11):
                cell.number_format = currency_fmt
            if col == 11:
                cell.font = green_font if (v or 0) >= 0 else red_font

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def generate_pdf_report(report: Dict) -> bytes:
    """Generate PDF report with tax calculations"""
    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    # Title
    pdf.set_fill_color(26, 54, 93)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font('Helvetica', 'B', 16)
    pdf.cell(0, 12, f'Capital Gains Tax Report - FY {report["fy_year"]}', ln=True, fill=True, align='C')
    pdf.ln(3)

    pdf.set_font('Helvetica', '', 8)
    pdf.set_text_color(100, 100, 100)
    pdf.cell(0, 5, f'Generated: {datetime.now().strftime("%d %b %Y, %I:%M %p")} | AI Trading Bot', ln=True, align='C')
    pdf.ln(8)

    # Trading Summary
    _pdf_section(pdf, 'TRADING SUMMARY')
    _pdf_row(pdf, 'Total Trades', str(report['total_trades']))
    _pdf_row(pdf, 'Profitable Trades', str(report['profitable_trades']))
    _pdf_row(pdf, 'Loss-Making Trades', str(report['loss_trades']))
    _pdf_row(pdf, 'Win Rate', f"{report['win_rate']}%")
    pdf.ln(5)

    # Financials
    _pdf_section(pdf, 'FINANCIAL SUMMARY')
    _pdf_row(pdf, 'Total Buy Value', f"{report['total_buy_value']:,.2f}")
    _pdf_row(pdf, 'Total Sell Value', f"{report['total_sell_value']:,.2f}")
    _pdf_row_colored(pdf, 'Total Profit', f"+{report['total_profit']:,.2f}", (22, 163, 74))
    _pdf_row_colored(pdf, 'Total Loss', f"-{report['total_loss']:,.2f}", (220, 38, 38))
    pnl = report['net_pnl']
    color = (22, 163, 74) if pnl >= 0 else (220, 38, 38)
    _pdf_row_colored(pdf, 'Net P&L', f"{'+'if pnl>=0 else ''}{pnl:,.2f}", color)
    _pdf_row(pdf, 'F&O Turnover', f"{report['turnover']:,.2f}")
    _pdf_row(pdf, 'STT Paid (Approx)', f"{report['stt_paid']:,.2f}")
    pdf.ln(5)

    # Tax Calculation
    _pdf_section(pdf, 'TAX CALCULATION')
    _pdf_row(pdf, 'STCG Tax @15%', f"{report['stcg_tax']:,.2f}")
    _pdf_row(pdf, 'Health & Education Cess @4%', f"{report['cess']:,.2f}")
    pdf.set_font('Helvetica', 'B', 11)
    pdf.set_text_color(220, 38, 38)
    pdf.cell(95, 8, 'Total Tax Liability', border='TB')
    pdf.cell(95, 8, f"{report['total_tax_liability']:,.2f}", border='TB', align='R')
    pdf.ln(8)
    pdf.set_text_color(0, 0, 0)
    _pdf_row(pdf, 'Effective Tax Rate', f"{report['effective_tax_rate']}%")
    pdf.ln(5)

    # Compliance
    _pdf_section(pdf, 'COMPLIANCE')
    _pdf_row(pdf, 'Tax Audit Required (Section 44AB)', 'Yes' if report['audit_required'] else 'No')
    _pdf_row(pdf, 'Audit Turnover Threshold', f"{report['audit_limit']:,.0f}")
    pdf.ln(8)

    # Monthly Breakdown Table
    if report['monthly_breakdown']:
        pdf.add_page()
        _pdf_section(pdf, 'MONTHLY BREAKDOWN')
        pdf.ln(3)

        # Table header
        col_widths = [25, 18, 30, 30, 30, 30, 27]
        headers = ['Month', 'Trades', 'Profit', 'Loss', 'Net P&L', 'Turnover', 'Tax']
        pdf.set_font('Helvetica', 'B', 8)
        pdf.set_fill_color(26, 54, 93)
        pdf.set_text_color(255, 255, 255)
        for w, h in zip(col_widths, headers):
            pdf.cell(w, 7, h, border=1, fill=True, align='C')
        pdf.ln()

        pdf.set_font('Helvetica', '', 8)
        for month, d in report['monthly_breakdown'].items():
            pdf.set_text_color(0, 0, 0)
            pdf.cell(col_widths[0], 6, month, border=1, align='C')
            pdf.cell(col_widths[1], 6, str(d['trades']), border=1, align='C')
            pdf.set_text_color(22, 163, 74)
            pdf.cell(col_widths[2], 6, f"{d['profit']:,.0f}", border=1, align='R')
            pdf.set_text_color(220, 38, 38)
            pdf.cell(col_widths[3], 6, f"{d['loss']:,.0f}", border=1, align='R')
            c = (22, 163, 74) if d['net_pnl'] >= 0 else (220, 38, 38)
            pdf.set_text_color(*c)
            pdf.cell(col_widths[4], 6, f"{d['net_pnl']:,.0f}", border=1, align='R')
            pdf.set_text_color(0, 0, 0)
            pdf.cell(col_widths[5], 6, f"{d['turnover']:,.0f}", border=1, align='R')
            pdf.cell(col_widths[6], 6, f"{d['total_tax']:,.0f}", border=1, align='R')
            pdf.ln()

    # Disclaimer
    pdf.ln(10)
    pdf.set_font('Helvetica', 'I', 7)
    pdf.set_text_color(120, 120, 120)
    pdf.multi_cell(0, 4, 'Disclaimer: This report is auto-generated for informational purposes only. Please consult a qualified CA/Tax Professional for actual tax filing. F&O income is treated as non-speculative business income under Indian tax law. The STCG @15% shown here is indicative - actual tax may vary based on your income slab and other factors. ITR-3 form is applicable for F&O traders.')

    return bytes(pdf.output())


def _pdf_section(pdf, title):
    pdf.set_font('Helvetica', 'B', 10)
    pdf.set_fill_color(226, 232, 240)
    pdf.set_text_color(0, 0, 0)
    pdf.cell(0, 8, f'  {title}', ln=True, fill=True)
    pdf.ln(2)

def _pdf_row(pdf, label, value):
    pdf.set_font('Helvetica', '', 9)
    pdf.set_text_color(60, 60, 60)
    pdf.cell(95, 7, f'  {label}')
    pdf.set_text_color(0, 0, 0)
    pdf.set_font('Helvetica', 'B', 9)
    pdf.cell(95, 7, value, align='R')
    pdf.ln()

def _pdf_row_colored(pdf, label, value, color):
    pdf.set_font('Helvetica', '', 9)
    pdf.set_text_color(60, 60, 60)
    pdf.cell(95, 7, f'  {label}')
    pdf.set_text_color(*color)
    pdf.set_font('Helvetica', 'B', 9)
    pdf.cell(95, 7, value, align='R')
    pdf.ln()
